from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from .models import Annual_data, Prov_Annual_data
from openpyxl.compat import range

def import_user(self, request, obj, change):

    wb = load_workbook(filename=obj.YDUserFile.path)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['version', 'attr', 'value', 'addr']
    lists = []
    users = request.user
    for row in range(2, 5):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        # for header in headers:
        revision = cell['version']
        prop = cell['attr']
        value = cell['value']
        repo = cell['addr']
        sql = KNSVNHistory(revision=revision, prop=prop, value=value, repo=repo, editor=users)
        sqllist.append(sql)
    KNSVNHistory.objects.bulk_create(sqllist)