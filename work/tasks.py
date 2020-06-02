from . import models
from openpyxl import load_workbook
import os
from django.core.exceptions import ValidationError


def parse_excel_and_save(fp: str, data_province_name: str, data_year: int):
    if not isinstance(fp, str):
        raise TypeError

    if not os.path.exists(fp):
        raise FileNotFoundError
    if not fp.lower().endswith('.xlsx'):
        raise ValueError
    try:
        r = models.import_excel_validator(fp)
    except ValidationError:
        raise
    pdr = models.ProvinceDataRecord.objects.filter(province__name=data_province_name, year=data_year).first()
    wb = load_workbook(fp, read_only=True)
    ws = wb[wb.sheetnames[0]]
    r = 0
    for row in ws.iter_rows():
        r += 1
        if r == 1:
            continue
        row = [c.value for c in row]
        _id, prov, area, year, r_d, per_unit_gdp, rural_urban, urban_per, rural_per, garbage,\
        bus_per, urban_sewage, edu_years, mortality, pension_cov, medical_cov, unemployment_cov, \
        pm25, so2_emissions, co2_per_gdp, cod_emissions, nh_emissions, water_per, water_per_gdp,\
        planting_area, ef_per = row
        if data_province_name != prov or year != data_year:
            print(f'Error {fp}, 第 {r} 行不是{data_province_name}{data_year}年度的数据')
            continue
        area_obj = models.Area.objects.filter(province__name=prov, name=area).first()
        print(prov, area, year)
        if not isinstance(area_obj, models.Area):
            print(f'Error {fp}, 第 {r} 行数据地区{area}不存在')
            continue
        city_data_record = models.CityDataRecord.objects.filter(
            area=area_obj,
            year=year
            ).first()
        if isinstance(city_data_record, models.CityDataRecord):
            # city_data_record.province_data = pdr
            city_data_record.r_d = r_d
            city_data_record.per_unit_gdp = per_unit_gdp
            city_data_record.rural_urban = rural_urban
            city_data_record.urban_per = urban_per
            city_data_record.rural_per = rural_per
            city_data_record.garbage = garbage
            city_data_record.bus_per = bus_per
            city_data_record.urban_sewage = urban_sewage
            city_data_record.edu_years = edu_years
            city_data_record.mortality = mortality
            city_data_record.pension_cov = pension_cov
            city_data_record.medical_cov = medical_cov
            city_data_record.unemployment_cov = unemployment_cov
            city_data_record.pm25 = pm25
            city_data_record.so2_emissions = so2_emissions
            city_data_record.co2_per_gdp = co2_per_gdp
            city_data_record.cod_emissions = cod_emissions
            city_data_record.nh_emissions = nh_emissions
            city_data_record.water_per = water_per
            city_data_record.water_per_gdp = water_per_gdp
            city_data_record.planting_area = planting_area
            city_data_record.ef_per = ef_per
            city_data_record.save()
        else:
            city_data_record = models.CityDataRecord.objects.create(
                # province_data=pdr,
                area=area_obj,
                year=year, r_d=r_d, per_unit_gdp=per_unit_gdp,
                rural_urban=rural_urban, urban_per=urban_per, rural_per=rural_per, garbage=garbage, bus_per=bus_per,
                urban_sewage=urban_sewage, edu_years=edu_years, mortality=mortality, pension_cov=pension_cov,
                medical_cov=medical_cov, unemployment_cov=unemployment_cov, pm25=pm25, so2_emissions=so2_emissions,
                co2_per_gdp=co2_per_gdp, cod_emissions=cod_emissions, nh_emissions=nh_emissions, water_per=water_per,
                water_per_gdp=water_per_gdp, planting_area=planting_area, ef_per=ef_per,
                )


city_keys = [
    'r_d', 'per_unit_gdp', 'rural_urban', 'urban_per', 'rural_per', 'garbage', 'bus_per', 'urban_sewage',
    'edu_years', 'mortality', 'pension_cov', 'medical_cov', 'unemployment_cov', 'pm25', 'so2_emissions',
    'co2_per_gdp', 'cod_emissions', 'nh_emissions', 'water_per', 'water_per_gdp', 'planting_area', 'ef_per'
    ]
province_keys = [
    'r_d',
    'renewable_energy_per',   # 省份有，地区无
    'per_unit_gdp', 'rural_urban', 'urban_per', 'rural_per', 'garbage', 'bus_per', 'urban_sewage',
    'edu_years', 'mortality', 'pension_cov', 'medical_cov', 'unemployment_cov', 'pm25', 'so2_emissions',
    'co2_per_gdp', 'cod_emissions', 'nh_emissions', 'water_per', 'water_per_gdp', 'planting_area', 'ef_per'
    ]

city_key_weight_method = {
    "r_d": "max",
    "per_unit_gdp": "min",
    "rural_urban": "max",
    "urban_per": "max",
    "rural_per": "max",
    "garbage": "max",
    "bus_per": "max",
    "urban_sewage": "max",
    "edu_years": "max",
    "mortality": "min",
    "pension_cov": "max",
    "medical_cov": "max",
    "unemployment_cov": "max",
    "pm25": "min",
    "so2_emissions": "min",
    "co2_per_gdp": "min",
    "cod_emissions": "min",
    "nh_emissions": "min",
    "water_per": "min",
    "water_per_gdp": "min",
    "planting_area": "max",
    "ef_per": "min"
    }


province_key_weight_method = {
    "r_d": "max",
    "renewable_energy_per": "max",  # 省份有，地区无
    "per_unit_gdp": "min",
    "rural_urban": "max",
    "urban_per": "max",
    "rural_per": "max",
    "garbage": "max",
    "bus_per": "max",
    "urban_sewage": "max",
    "edu_years": "max",
    "mortality": "min",
    "pension_cov": "max",
    "medical_cov": "max",
    "unemployment_cov": "max",
    "pm25": "min",
    "so2_emissions": "min",
    "co2_per_gdp": "min",
    "cod_emissions": "min",
    "nh_emissions": "min",
    "water_per": "min",
    "water_per_gdp": "min",
    "planting_area": "max",
    "ef_per": "min"
    }

city_calc_map = {
    'green_innovation': ['r_d'],
    'energy_use': ['per_unit_gdp'],
    'parma_ratio': ['rural_urban'],
    'income': ['urban_per', 'rural_per'],
    'infrastructure': ["garbage", "bus_per", "urban_sewage"],
    'education': ['edu_years'],
    'life_expectancy': ['mortality'],
    'social_security': ["pension_cov", "medical_cov", "unemployment_cov"],
    'air_pollution': ["pm25", "so2_emissions"],
    'greenhouse': ['co2_per_gdp'],
    'nitrogen': ["cod_emissions", "nh_emissions"],
    'water_withdrawal': ["water_per", "water_per_gdp"],
    "land_use": ["planting_area"],
    'EF': ['ef_per'],

    #
    'city_green_economy': ["green_innovation", "energy_use", "parma_ratio", "income", "infrastructure", "education", "life_expectancy", "social_security", "air_pollution"],
    'city_sustainable': ["greenhouse", "nitrogen", "water_withdrawal", "land_use", "EF"],

    #
    'city_gep_plus': ["city_green_economy", "city_sustainable"]
    }


province_calc_map = {
    'green_innovation': ['r_d'],
    'renewable_energy': ['renewable_energy_per'],  # 省份有，地区无
    'energy_use': ['per_unit_gdp'],
    'parma_ratio': ['rural_urban'],
    'income': ['urban_per', 'rural_per'],
    'infrastructure': ["garbage", "bus_per", "urban_sewage"],
    'education': ['edu_years'],
    'life_expectancy': ['mortality'],
    'social_security': ["pension_cov", "medical_cov", "unemployment_cov"],
    'air_pollution': ["pm25", "so2_emissions"],
    'greenhouse': ['co2_per_gdp'],
    'nitrogen': ["cod_emissions", "nh_emissions"],
    'water_withdrawal': ["water_per", "water_per_gdp"],
    "land_use": ["planting_area"],
    'EF': ['ef_per'],

    #
    'city_green_economy': ["green_innovation", "energy_use", "parma_ratio", "income", "infrastructure", "education", "life_expectancy", "social_security", "air_pollution"],
    'city_sustainable': ["greenhouse", "nitrogen", "water_withdrawal", "land_use", "EF"],

    #
    'city_gep_plus': ["city_green_economy", "city_sustainable"]
    }


def city_find_base(k: str, bs: list):
    if k not in city_calc_map:
        raise ValueError
    for i in city_calc_map[k]:
        if i in city_keys:
            bs.append(i)
            continue
        elif i in city_calc_map:
            city_find_base(i, bs)
        else:
            return


def province_find_base(k: str, bs: list):
    if k not in province_calc_map:
        raise ValueError
    for i in province_calc_map[k]:
        if i in province_keys:
            bs.append(i)
            continue
        elif i in province_calc_map:
            province_find_base(i, bs)
        else:
            return


def city_calc_gep(data: dict, key):
    bs = []
    city_find_base(key, bs)
    if not bs:
        raise ValueError
    elif len(bs) == 1:
        return data[f'{bs[0]}_gep']
    else:
        return sum([data[f'{k}_gep'] * data[f'{k}_weight'] for k in bs]) / sum([data[f'{k}_weight'] for k in bs])


def province_calc_gep(data: dict, key):
    bs = []
    province_find_base(key, bs)
    print(key, bs)
    if not bs:
        raise ValueError
    elif len(bs) == 1:
        return data[f'{bs[0]}_gep']
    else:
        return sum([data[f'{k}_gep'] * data[f'{k}_weight'] for k in bs]) / sum([data[f'{k}_weight'] for k in bs])


# def __calc_diff(key: str, current: (float, int), last: (float, int)):
#     if not isinstance(key, str):
#         raise TypeError
#     if not isinstance(current, (float, int)):
#         raise TypeError
#     if not isinstance(last, (float, int)):
#         raise TypeError
#     if key not in city_key_weight_method:
#         raise ValueError
#     if current == 0:
#         raise ValueError
#     if last == 0:
#         raise ValueError
#
#     diff = current - last
#     if city_key_weight_method[key] == 'min':
#         return -diff
#     else:
#         return diff


def __calc_step(last: (float, int), current: (float, int), target: (float, int)):
    if target - last == 0:
        return 1
    if current - last == 0:
        return 0
    return (current - last) / (target - last)


def __get_city_target(data: dict, key: str):
    if key not in city_key_weight_method:
        raise ValueError
    if not isinstance(data, dict):
        raise TypeError
    k = f'{key}_{city_key_weight_method[key]}'
    if k not in data:
        raise ValueError
    v = data[k]
    if not isinstance(v, (float, int)):
        raise ValueError
    if v == 0:
        raise ValueError
    return v


def city_calc_weight(target: (float, int), last: (float, int), key: str):
    if key not in city_key_weight_method:
        raise ValueError
    if not isinstance(target, (int, float)):
        raise TypeError
    if not isinstance(last, (float, int)):
        raise TypeError
    if target == 0:
        raise ValueError
    if last == 0:
        raise ValueError
    if city_key_weight_method[key] == 'max':
        return target / last
    else:
        return last / target


def province_calc_weight(target: (float, int), last: (float, int), key: str):
    if key not in province_key_weight_method:
        raise ValueError
    if not isinstance(target, (int, float)):
        raise TypeError
    if not isinstance(last, (float, int)):
        raise TypeError
    if target == 0:
        raise ValueError
    if last == 0:
        raise ValueError
    if province_key_weight_method[key] == 'max':
        return target / last
    else:
        return last / target


def __get_province_target(p: models.ProvinceDataRecord, key: str):
    if not isinstance(p, models.ProvinceDataRecord):
        raise TypeError
    if not isinstance(key, str):
        raise TypeError
    if key not in province_keys:
        raise ValueError
    k = f'{key}_target'
    if k not in p.__dict__:
        raise ValueError
    return p.__dict__[k]


def do_gather_data(province_name: str, year: int):
    if not isinstance(province_name, str):
        raise TypeError('省份名称（province_name）应位字符型数据')
    if not isinstance(year, int):
        raise TypeError('年份（year）应该是4位正整数')
    if not province_name:
        raise ValidationError
    province_data_record_of_this_year = models.ProvinceDataRecord.objects.filter(province__name=province_name, year=year).first()
    # if not isinstance(province_data_record_of_this_year, models.ProvinceDataRecord):
    #     return f'{province_name}-{year}的数据不存在'
    province_data_record_of_last_year = models.ProvinceDataRecord.objects.filter(province__name=province_name, year=year-1).first()
    # if not isinstance(province_data_record_of_last_year, models.ProvinceDataRecord):
    #     return f'{province_name}-{year-1}的数据不存在'
    cities_of_this_year = [
        city_data_record for city_data_record in models.CityDataRecord.objects.filter(
            area__province__name=province_name, year=year).all()
        ]

    cities_of_last_year = [
        city_data_record for city_data_record in models.CityDataRecord.objects.filter(
            area__province__name=province_name, year=year-1).all()
        ]

    cities_of_this_year_max_min_dict = {
        }

    cites_num = models.Area.objects.filter(province__name=province_name).count()

    for k in city_keys:
        values = [city_data_record.__dict__[k] for city_data_record in cities_of_this_year]
        if values:
            cities_of_this_year_max_min_dict[f'{k}_min'] = min(values)
            cities_of_this_year_max_min_dict[f'{k}_max'] = max(values)

    if not province_data_record_of_this_year \
            or not province_data_record_of_last_year \
            or not cities_of_this_year\
            or not cities_of_last_year \
            or len(cities_of_last_year) != cites_num\
            or len(cities_of_this_year) != cites_num\
            or len(cities_of_this_year_max_min_dict) != len(city_keys)*2:
        tag = False
    else:
        tag = True

    return tag, province_data_record_of_this_year, province_data_record_of_last_year, cities_of_this_year,\
        cities_of_this_year_max_min_dict, cities_of_last_year


def do_calc(province_name: str, year: int):
    tag, province_data_record_of_this_year, province_data_record_of_last_year, cities_of_this_year, \
        cities_of_this_year_max_min_dict, cities_of_last_year = do_gather_data(province_name=province_name, year=year)
    if not tag:
        raise ValueError(f'数据不足不能计算{province_name}{year}')
    province_data_record_of_this_year.status = 3
    province_data_record_of_this_year.save()
    # 计算省份
    # try:
    #     calc_province_data_record(
    #         province_data_record=province_data_record_of_this_year,
    #         province_data_record_last_year=province_data_record_of_last_year)
    # except ZeroDivisionError:
    #     print(f'{province_name}-{year}, 计算遇到错误，除数为0')
    #     raise

    new_calc_province_addition(
        province_data_record=province_data_record_of_this_year,
        province_data_record_last_year=province_data_record_of_last_year)

    # 计算城市
    for city_data_record in cities_of_this_year:
        city_data_record_last_year = ''
        for c in cities_of_last_year:
            if c.area.name == city_data_record.area.name:
                city_data_record_last_year = c
                break
        if not isinstance(city_data_record_last_year, models.CityDataRecord):
            continue
        # try:
        #     calc_city_data_record(
        #         city_data_record=city_data_record,
        #         city_data_record_last_year=city_data_record_last_year,
        #         cities_of_this_year_max_min_dict=cities_of_this_year_max_min_dict
        #         )
        # except ZeroDivisionError:
        #     print(f'{province_name}-{city_data_record.area.name}-{year}, 计算遇到错误，除数为0')
        #     continue
        new_calc_city_addition(
            city_data_record=city_data_record,
            city_data_record_last_year=city_data_record_last_year,
            cities_of_this_year_max_min_dict=cities_of_this_year_max_min_dict
            )
    province_data_record_of_this_year.status = 5
    province_data_record_of_this_year.save()


def new_calc_city_addition(
        city_data_record: models.CityDataRecord, city_data_record_last_year: models.CityDataRecord,
        cities_of_this_year_max_min_dict: dict):
    data = {}
    for k in city_keys:
        last = city_data_record_last_year.__dict__[k]
        current = city_data_record.__dict__[k]
        target = __get_city_target(data=cities_of_this_year_max_min_dict, key=k)
        # diff = __calc_diff(key=k, current=current, last=last)
        weight = city_calc_weight(target=target, last=last, key=k)
        step = __calc_step(last=last, current=current, target=target)
        gep = step * weight
        data[f'{k}_weight'] = weight
        data[f'{k}_gep'] = gep

    for k in city_calc_map:
        city_data_record.__dict__[k] = city_calc_gep(data, k)

    city_data_record.save()


def new_calc_province_addition(
        province_data_record: models.ProvinceDataRecord,
        province_data_record_last_year: models.ProvinceDataRecord,):
    data = {}
    for k in province_keys:
        last = province_data_record_last_year.__dict__[k]
        current = province_data_record.__dict__[k]
        target = __get_province_target(p=province_data_record, key=k)
        # diff = __calc_diff(key=k, current=current, last=last)
        weight = province_calc_weight(target=target, last=last, key=k)
        step = __calc_step(last=last, current=current, target=target)
        gep = step * weight
        print(f'{k}, weight: {weight}, step: {step}, gep: {gep}')
        data[f'{k}_weight'] = weight
        data[f'{k}_gep'] = gep

    for k in province_calc_map:
        v = province_calc_gep(data, k)
        province_data_record.__dict__[k] = v
        print(k, v)

    province_data_record.save()



# def calc_city_data_record(city_data_record: models.CityDataRecord, city_data_record_last_year: models.CityDataRecord,
#                           cities_of_this_year_max_min_dict: dict):
#     # 绿色创新
#     r_d_last = city_data_record_last_year.r_d
#     r_d_target = cities_of_this_year_max_min_dict['r_d_max']
#     r_d_current = city_data_record.r_d
#     r_d_weight = r_d_target / r_d_last
#     r_d_diff = r_d_current - r_d_last
#     r_d_run = r_d_target - r_d_last
#     if r_d_diff == 0:
#         r_d_step = 0
#     elif r_d_run == 0:
#         r_d_step = 1
#     else:
#         r_d_step = r_d_diff / r_d_run
#     r_d_gep = r_d_step * r_d_weight
#
#     green_innovation = r_d_gep  # 其余各项都是按照这个规则，
#     # 两个单项指标得分*权重之和/两个单项指标权重之和 —————— 就是单项的gep分值相加的和，除以各单项的权重的和
#
#     # 能源利用
#     per_unit_gdp_last = city_data_record_last_year.per_unit_gdp
#     per_unit_gdp_target = cities_of_this_year_max_min_dict['per_unit_gdp_min']
#     per_unit_gdp_current = city_data_record.per_unit_gdp
#     per_unit_gdp_weight = per_unit_gdp_last / per_unit_gdp_target
#
#     energy_use = cities_of_this_year_max_min_dict['per_unit_gdp_min']
#     # 帕尔玛比率
#     rural_urban_last = city_data_record_last_year.rural_urban
#     rural_urban_target = cities_of_this_year_max_min_dict['rural_urban_max']
#     rural_urban_current = city_data_record.rural_urban
#     rural_urban_weight = rural_urban_target / rural_urban_last
#
#     parma_ratio = cities_of_this_year_max_min_dict['rural_urban_max']
#
#     # 收入
#     urban_per_last = city_data_record_last_year.urban_per
#     urban_per_target = cities_of_this_year_max_min_dict['urban_per_max']
#     urban_per_current = city_data_record.urban_per
#     urban_per_weight = urban_per_target / urban_per_last
#
#     rural_per_last = city_data_record_last_year.rural_per
#     rural_per_target = cities_of_this_year_max_min_dict['rural_per_max']
#     rural_per_current = city_data_record.rural_per
#     rural_per_weight = rural_per_target / rural_per_last
#
#     income = (urban_per_current * urban_per_weight + rural_per_current * rural_per_weight) / (
#                 urban_per_weight + rural_per_weight)
#
#     # 基础设施建设
#     garbage_last = city_data_record_last_year.garbage
#     garbage_target = cities_of_this_year_max_min_dict['garbage_max']
#     garbage_current = city_data_record.garbage
#     garbage_weight = garbage_target / garbage_last
#
#     bus_per_last = city_data_record_last_year.bus_per
#     bus_per_target = cities_of_this_year_max_min_dict['bus_per_max']
#     bus_per_current = city_data_record.bus_per
#     bus_per_weight = bus_per_target / bus_per_last
#
#     urban_sewage_last = city_data_record_last_year.urban_sewage
#     urban_sewage_target = cities_of_this_year_max_min_dict['urban_sewage_max']
#     urban_sewage_current = city_data_record.urban_sewage
#     urban_sewage_weight = urban_sewage_target / urban_sewage_last
#
#     infrastructure = (
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight
#                      ) / (
#         garbage_weight + bus_per_weight + urban_sewage_weight)
#
#     # 教育
#     edu_years_last = city_data_record_last_year.edu_years
#     edu_years_target = cities_of_this_year_max_min_dict['edu_years_max']
#     edu_years_current = city_data_record.edu_years
#     edu_years_weight = edu_years_target / edu_years_last
#
#     education = cities_of_this_year_max_min_dict['edu_years_max']
#
#     # 预期寿命
#     mortality_last = city_data_record_last_year.mortality
#     mortality_target = cities_of_this_year_max_min_dict['mortality_min']
#     mortality_current = city_data_record.mortality
#     mortality_weight = mortality_last / mortality_target
#
#     life_expectancy = cities_of_this_year_max_min_dict['mortality_min']
#
#     # 社会保障
#     pension_cov_last = city_data_record_last_year.pension_cov
#     pension_cov_target = cities_of_this_year_max_min_dict['pension_cov_max']
#     pension_cov_current = city_data_record.pension_cov
#     pension_cov_weight = pension_cov_target / pension_cov_last
#
#     medical_cov_last = city_data_record_last_year.medical_cov
#     medical_cov_target = cities_of_this_year_max_min_dict['medical_cov_max']
#     medical_cov_current = city_data_record.medical_cov
#     medical_cov_weight = medical_cov_target / medical_cov_last
#
#     unemployment_cov_last = city_data_record_last_year.unemployment_cov
#     unemployment_cov_target = cities_of_this_year_max_min_dict['unemployment_cov_max']
#     unemployment_cov_current = city_data_record.unemployment_cov
#     unemployment_cov_weight = unemployment_cov_target / unemployment_cov_last
#
#     social_security = (
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight
#                       ) / (
#         pension_cov_weight +
#         medical_cov_weight +
#         unemployment_cov_weight)
#
#     # 大气污染
#     pm25_last = city_data_record_last_year.pm25
#     pm25_target = cities_of_this_year_max_min_dict['pm25_min']
#     pm25_current = city_data_record.pm25
#     pm25_weight = pm25_last / pm25_target
#
#     so2_emissions_last = city_data_record_last_year.so2_emissions
#     so2_emissions_target = cities_of_this_year_max_min_dict['so2_emissions_min']
#     so2_emissions_current = city_data_record.so2_emissions
#     so2_emissions_weight = so2_emissions_last/so2_emissions_target
#
#     air_pollution = (
#         pm25_current * pm25_weight + so2_emissions_current * so2_emissions_weight) / (
#         pm25_weight + so2_emissions_weight)
#
#     # 温室气体排放
#     co2_per_gdp_last = city_data_record_last_year.co2_per_gdp
#     co2_per_gdp_target = cities_of_this_year_max_min_dict['co2_per_gdp_min']
#     co2_per_gdp_current = city_data_record.co2_per_gdp
#     co2_per_gdp_weight = co2_per_gdp_last / co2_per_gdp_target
#
#     greenhouse = cities_of_this_year_max_min_dict['co2_per_gdp_min']
#
#     # 氮排放
#     cod_emissions_last = city_data_record_last_year.cod_emissions
#     cod_emissions_target = cities_of_this_year_max_min_dict['cod_emissions_min']
#     cod_emissions_current = city_data_record.cod_emissions
#     cod_emissions_weight = cod_emissions_last / cod_emissions_target
#
#     nh_emissions_last = city_data_record_last_year.nh_emissions
#     nh_emissions_target = cities_of_this_year_max_min_dict['nh_emissions_min']
#     nh_emissions_current = city_data_record.nh_emissions
#     nh_emissions_weight = nh_emissions_last / nh_emissions_target
#
#     nitrogen = (cod_emissions_current * cod_emissions_weight + nh_emissions_current * nh_emissions_weight) / (
#                 cod_emissions_weight + nh_emissions_weight)
#
#     # 取水量
#     water_per_last = city_data_record_last_year.water_per
#     water_per_target = cities_of_this_year_max_min_dict['water_per_min']
#     water_per_current = city_data_record.water_per
#     water_per_weight = water_per_last / water_per_target
#
#     water_per_gdp_last = city_data_record_last_year.water_per_gdp
#     water_per_gdp_target = cities_of_this_year_max_min_dict['water_per_gdp_min']
#     water_per_gdp_current = city_data_record.water_per_gdp
#     water_per_gdp_weight = water_per_gdp_last / water_per_gdp_target
#
#     water_withdrawal = (water_per_current * water_per_weight + water_per_gdp_current * water_per_gdp_weight) / (
#                 water_per_weight + water_per_gdp_weight)
#
#     # 土地利用
#     planting_area_last = city_data_record_last_year.planting_area
#     planting_area_target = cities_of_this_year_max_min_dict['planting_area_max']
#     planting_area_current = city_data_record.planting_area
#     planting_area_weight = planting_area_target / planting_area_last
#
#     land_use = cities_of_this_year_max_min_dict['planting_area_max']
#
#     # 生态足迹
#     ef_per_last = city_data_record_last_year.ef_per
#     ef_per_target = cities_of_this_year_max_min_dict['ef_per_min']
#     ef_per_current = city_data_record.ef_per
#     ef_per_weight = ef_per_last / ef_per_target
#
#     EF = cities_of_this_year_max_min_dict['ef_per_min']
#
#     # 绿色经济
#     city_green_economy = (
#         r_d_current * r_d_weight +
#         per_unit_gdp_current * per_unit_gdp_weight +
#         rural_urban_current * rural_urban_weight +
#         urban_per_current * urban_per_weight +
#         rural_per_current * rural_per_weight +
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight +
#         edu_years_current * edu_years_weight +
#         mortality_current * mortality_weight +
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight +
#         pm25_current * pm25_weight +
#         so2_emissions_current * so2_emissions_weight
#                          ) / (
#         r_d_weight + per_unit_gdp_weight + rural_urban_weight + urban_per_weight + rural_per_weight +
#         garbage_weight + bus_per_weight + urban_sewage_weight + edu_years_weight + mortality_weight +
#         pension_cov_weight + medical_cov_weight + unemployment_cov_weight + pm25_weight +
#         so2_emissions_weight
#
#                          )
#
#     # 可持续发展
#     city_sustainable = (
#         co2_per_gdp_current * co2_per_gdp_weight +
#         cod_emissions_current * cod_emissions_weight +
#         nh_emissions_current * nh_emissions_weight +
#         water_per_current * water_per_weight +
#         water_per_gdp_current * water_per_gdp_weight +
#         planting_area_current * planting_area_weight +
#         ef_per_current * ef_per_weight
#                        ) / (
#         co2_per_gdp_weight +
#         cod_emissions_weight +
#         nh_emissions_weight +
#         water_per_weight +
#         water_per_gdp_weight +
#         planting_area_weight +
#         ef_per_weight
#     )
#
#     # GEP+
#     city_gep_plus = (
#         # 绿色经济
#         r_d_current * r_d_weight +
#         per_unit_gdp_current * per_unit_gdp_weight +
#         rural_urban_current * rural_urban_weight +
#         urban_per_current * urban_per_weight +
#         rural_per_current * rural_per_weight +
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight +
#         edu_years_current * edu_years_weight +
#         mortality_current * mortality_weight +
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight +
#         pm25_current * pm25_weight +
#         so2_emissions_current * so2_emissions_weight +
#         # 可持续发展
#         co2_per_gdp_current * co2_per_gdp_weight +
#         cod_emissions_current * cod_emissions_weight +
#         nh_emissions_current * nh_emissions_weight +
#         water_per_current * water_per_weight +
#         water_per_gdp_current * water_per_gdp_weight +
#         planting_area_current * planting_area_weight +
#         ef_per_current * ef_per_weight
#                     ) / (
#         r_d_weight +
#         per_unit_gdp_weight +
#         rural_urban_weight +
#         urban_per_weight +
#         rural_per_weight +
#         garbage_weight +
#         bus_per_weight +
#         urban_sewage_weight +
#         edu_years_weight +
#         mortality_weight +
#         pension_cov_weight +
#         medical_cov_weight +
#         unemployment_cov_weight +
#         pm25_weight +
#         so2_emissions_weight +
#         co2_per_gdp_weight +
#         cod_emissions_weight +
#         nh_emissions_weight +
#         water_per_weight +
#         water_per_gdp_weight +
#         planting_area_weight +
#         ef_per_weight)
#
#     city_data_record.green_innovation = green_innovation
#     city_data_record.energy_use = energy_use
#     city_data_record.parma_ratio = parma_ratio
#     city_data_record.income = income
#     city_data_record.infrastructure = infrastructure
#     city_data_record.education = education
#     city_data_record.life_expectancy = life_expectancy
#     city_data_record.social_security = social_security
#     city_data_record.air_pollution = air_pollution
#     city_data_record.greenhouse = greenhouse
#     city_data_record.nitrogen = nitrogen
#     city_data_record.water_withdrawal = water_withdrawal
#     city_data_record.land_use = land_use
#     city_data_record.EF = EF
#     city_data_record.city_green_economy = city_green_economy
#     city_data_record.city_sustainable = city_sustainable
#     city_data_record.city_gep_plus = city_gep_plus
#
#     city_data_record.save()
#
#
# def calc_province_data_record(
#         province_data_record: models.ProvinceDataRecord,
#         province_data_record_last_year: models.ProvinceDataRecord,):
#     # 绿色创新
#     r_d_last = province_data_record_last_year.r_d
#     r_d_target = province_data_record.r_d_target  # province_data_record.r_d_target
#     r_d_current = province_data_record.r_d
#     r_d_weight = r_d_target / r_d_last
#
#     green_innovation = province_data_record.r_d_target
#
#     # 能源利用
#     per_unit_gdp_last = province_data_record_last_year.per_unit_gdp
#     per_unit_gdp_target = province_data_record.per_unit_gdp_target
#     per_unit_gdp_current = province_data_record.per_unit_gdp
#     per_unit_gdp_weight = per_unit_gdp_last / per_unit_gdp_target
#
#     energy_use = province_data_record.per_unit_gdp_target
#     # 帕尔玛比率
#     rural_urban_last = province_data_record_last_year.rural_urban
#     rural_urban_target = province_data_record.rural_urban_target
#     rural_urban_current = province_data_record.rural_urban
#     rural_urban_weight = rural_urban_target / rural_urban_last
#
#     parma_ratio = province_data_record.rural_urban_target
#
#     # 收入
#     urban_per_last = province_data_record_last_year.urban_per
#     urban_per_target = province_data_record.urban_per_target
#     urban_per_current = province_data_record.urban_per
#     urban_per_weight = urban_per_target / urban_per_last
#
#     rural_per_last = province_data_record_last_year.rural_per
#     rural_per_target = province_data_record.rural_per_target
#     rural_per_current = province_data_record.rural_per
#     rural_per_weight = rural_per_target / rural_per_last
#
#     income = (urban_per_current * urban_per_weight + rural_per_current * rural_per_weight) / (
#                 urban_per_weight + rural_per_weight)
#
#     # 基础设施建设
#     garbage_last = province_data_record_last_year.garbage
#     garbage_target = province_data_record.garbage_target
#     garbage_current = province_data_record.garbage
#     garbage_weight = garbage_target / garbage_last
#
#     bus_per_last = province_data_record_last_year.bus_per
#     bus_per_target = province_data_record.bus_per_target
#     bus_per_current = province_data_record.bus_per
#     bus_per_weight = bus_per_target / bus_per_last
#
#     urban_sewage_last = province_data_record_last_year.urban_sewage
#     urban_sewage_target = province_data_record.urban_sewage_target
#     urban_sewage_current = province_data_record.urban_sewage
#     urban_sewage_weight = urban_sewage_target / urban_sewage_last
#
#     infrastructure = (
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight
#                      ) / (
#         garbage_weight + bus_per_weight + urban_sewage_weight)
#
#     # 教育
#     edu_years_last = province_data_record_last_year.edu_years
#     edu_years_target = province_data_record.edu_years_target
#     edu_years_current = province_data_record.edu_years
#     edu_years_weight = edu_years_target / edu_years_last
#
#     education = province_data_record.edu_years_target
#
#     # 预期寿命
#     mortality_last = province_data_record_last_year.mortality
#     mortality_target = province_data_record.mortality_target
#     mortality_current = province_data_record.mortality
#     mortality_weight = mortality_last / mortality_target
#
#     life_expectancy = province_data_record.mortality_target
#
#     # 社会保障
#     pension_cov_last = province_data_record_last_year.pension_cov
#     pension_cov_target = province_data_record.pension_cov_target
#     pension_cov_current = province_data_record.pension_cov
#     pension_cov_weight = pension_cov_target / pension_cov_last
#
#     medical_cov_last = province_data_record_last_year.medical_cov
#     medical_cov_target = province_data_record.medical_cov_target
#     medical_cov_current = province_data_record.medical_cov
#     medical_cov_weight = medical_cov_target / medical_cov_last
#
#     unemployment_cov_last = province_data_record_last_year.unemployment_cov
#     unemployment_cov_target = province_data_record.unemployment_cov_target
#     unemployment_cov_current = province_data_record.unemployment_cov
#     unemployment_cov_weight = unemployment_cov_target / unemployment_cov_last
#
#     social_security = (
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight
#                       ) / (
#         pension_cov_weight +
#         medical_cov_weight +
#         unemployment_cov_weight)
#
#     # 大气污染
#     pm25_last = province_data_record_last_year.pm25
#     pm25_target = province_data_record.pm25_target
#     pm25_current = province_data_record.pm25
#     pm25_weight = pm25_last / pm25_target
#
#     so2_emissions_last = province_data_record_last_year.so2_emissions
#     so2_emissions_target = province_data_record.so2_emissions_target
#     so2_emissions_current = province_data_record.so2_emissions
#     so2_emissions_weight = so2_emissions_last/so2_emissions_target
#
#     air_pollution = (
#         pm25_current * pm25_weight + so2_emissions_current * so2_emissions_weight) / (
#         pm25_weight + so2_emissions_weight)
#
#     # 温室气体排放
#     co2_per_gdp_last = province_data_record_last_year.co2_per_gdp
#     co2_per_gdp_target = province_data_record.co2_per_gdp_target
#     co2_per_gdp_current = province_data_record.co2_per_gdp
#     co2_per_gdp_weight = co2_per_gdp_last / co2_per_gdp_target
#
#     greenhouse = province_data_record.co2_per_gdp_target
#
#     # 氮排放
#     cod_emissions_last = province_data_record_last_year.cod_emissions
#     cod_emissions_target = province_data_record.cod_emissions_target
#     cod_emissions_current = province_data_record.cod_emissions
#     cod_emissions_weight = cod_emissions_last / cod_emissions_target
#
#     nh_emissions_last = province_data_record_last_year.nh_emissions
#     nh_emissions_target = province_data_record.nh_emissions_target
#     nh_emissions_current = province_data_record.nh_emissions
#     nh_emissions_weight = nh_emissions_last / nh_emissions_target
#
#     nitrogen = (cod_emissions_current * cod_emissions_weight + nh_emissions_current * nh_emissions_weight) / (
#                 cod_emissions_weight + nh_emissions_weight)
#
#     # 取水量
#     water_per_last = province_data_record_last_year.water_per
#     water_per_target = province_data_record.water_per_target
#     water_per_current = province_data_record.water_per
#     water_per_weight = water_per_last / water_per_target
#
#     water_per_gdp_last = province_data_record_last_year.water_per_gdp
#     water_per_gdp_target = province_data_record.water_per_gdp_target
#     water_per_gdp_current = province_data_record.water_per_gdp
#     water_per_gdp_weight = water_per_gdp_last / water_per_gdp_target
#
#     water_withdrawal = (water_per_current * water_per_weight + water_per_gdp_current * water_per_gdp_weight) / (
#                 water_per_weight + water_per_gdp_weight)
#
#     # 土地利用
#     planting_area_last = province_data_record_last_year.planting_area
#     planting_area_target = province_data_record.planting_area_target
#     planting_area_current = province_data_record.planting_area
#     planting_area_weight = planting_area_target / planting_area_last
#
#     land_use = province_data_record.planting_area_target
#
#     # 生态足迹
#     ef_per_last = province_data_record_last_year.ef_per
#     ef_per_target = province_data_record.ef_per_target
#     ef_per_current = province_data_record.ef_per
#     ef_per_weight = ef_per_last / ef_per_target
#
#     EF = province_data_record.ef_per_target
#
#     # 绿色经济
#     province_green_economy = (
#         r_d_current * r_d_weight +
#         per_unit_gdp_current * per_unit_gdp_weight +
#         rural_urban_current * rural_urban_weight +
#         urban_per_current * urban_per_weight +
#         rural_per_current * rural_per_weight +
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight +
#         edu_years_current * edu_years_weight +
#         mortality_current * mortality_weight +
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight +
#         pm25_current * pm25_weight +
#         so2_emissions_current * so2_emissions_weight
#                          ) / (
#         r_d_weight + per_unit_gdp_weight + rural_urban_weight + urban_per_weight + rural_per_weight +
#         garbage_weight + bus_per_weight + urban_sewage_weight + edu_years_weight + mortality_weight +
#         pension_cov_weight + medical_cov_weight + unemployment_cov_weight + pm25_weight +
#         so2_emissions_weight
#
#                          )
#
#     # 可持续发展
#     province_sustainable = (
#         co2_per_gdp_current * co2_per_gdp_weight +
#         cod_emissions_current * cod_emissions_weight +
#         nh_emissions_current * nh_emissions_weight +
#         water_per_current * water_per_weight +
#         water_per_gdp_current * water_per_gdp_weight +
#         planting_area_current * planting_area_weight +
#         ef_per_current * ef_per_weight
#                        ) / (
#         co2_per_gdp_weight +
#         cod_emissions_weight +
#         nh_emissions_weight +
#         water_per_weight +
#         water_per_gdp_weight +
#         planting_area_weight +
#         ef_per_weight
#     )
#
#     # GEP+
#     province_gep_plus = (
#         # 绿色经济
#         r_d_current * r_d_weight +
#         per_unit_gdp_current * per_unit_gdp_weight +
#         rural_urban_current * rural_urban_weight +
#         urban_per_current * urban_per_weight +
#         rural_per_current * rural_per_weight +
#         garbage_current * garbage_weight +
#         bus_per_current * bus_per_weight +
#         urban_sewage_current * urban_sewage_weight +
#         edu_years_current * edu_years_weight +
#         mortality_current * mortality_weight +
#         pension_cov_current * pension_cov_weight +
#         medical_cov_current * medical_cov_weight +
#         unemployment_cov_current * unemployment_cov_weight +
#         pm25_current * pm25_weight +
#         so2_emissions_current * so2_emissions_weight +
#         # 可持续发展
#         co2_per_gdp_current * co2_per_gdp_weight +
#         cod_emissions_current * cod_emissions_weight +
#         nh_emissions_current * nh_emissions_weight +
#         water_per_current * water_per_weight +
#         water_per_gdp_current * water_per_gdp_weight +
#         planting_area_current * planting_area_weight +
#         ef_per_current * ef_per_weight
#                     ) / (
#         r_d_weight +
#         per_unit_gdp_weight +
#         rural_urban_weight +
#         urban_per_weight +
#         rural_per_weight +
#         garbage_weight +
#         bus_per_weight +
#         urban_sewage_weight +
#         edu_years_weight +
#         mortality_weight +
#         pension_cov_weight +
#         medical_cov_weight +
#         unemployment_cov_weight +
#         pm25_weight +
#         so2_emissions_weight +
#         co2_per_gdp_weight +
#         cod_emissions_weight +
#         nh_emissions_weight +
#         water_per_weight +
#         water_per_gdp_weight +
#         planting_area_weight +
#         ef_per_weight)
#
#     province_data_record.green_innovation = green_innovation
#     province_data_record.energy_use = energy_use
#     province_data_record.parma_ratio = parma_ratio
#     province_data_record.income = income
#     province_data_record.infrastructure = infrastructure
#     province_data_record.education = education
#     province_data_record.life_expectancy = life_expectancy
#     province_data_record.social_security = social_security
#     province_data_record.air_pollution = air_pollution
#     province_data_record.greenhouse = greenhouse
#     province_data_record.nitrogen = nitrogen
#     province_data_record.water_withdrawal = water_withdrawal
#     province_data_record.land_use = land_use
#     province_data_record.EF = EF
#     province_data_record.city_green_economy = province_green_economy
#     province_data_record.city_sustainable = province_sustainable
#     province_data_record.city_gep_plus = province_gep_plus
#
#     province_data_record.save()

















        
        

        


