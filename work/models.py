from django.db import models
from django.core.exceptions import ValidationError
# Create your models here.
from openpyxl import load_workbook
from django.core.validators import FileExtensionValidator


class Attach(models.Model):
    name = models.CharField(verbose_name='附件名称', null=True, blank=True, max_length=128)
    attach = models.FileField(verbose_name='附件', null=False, blank=False, upload_to='attach/%Y-%m-%d')
    created = models.DateTimeField(verbose_name='上传时间', auto_now_add=True)
    updated = models.DateTimeField(verbose_name='修改时间', auto_now=True)

    article = models.ForeignKey(to='Article', on_delete=models.CASCADE, related_name='attach_ls')

    class Meta:
        ordering = '-created', 'name', 'attach'
        verbose_name = '附件'
        verbose_name_plural = verbose_name

    def __str__(self):
        return self.name or self.attach.name


class Article(models.Model):

    title = models.CharField(verbose_name='标题', max_length=255, db_index=True)
    author = models.CharField(verbose_name='作者', max_length=64, null=True, blank=True)
    info = models.CharField(verbose_name='摘要', max_length=255, null=True, blank=True)
    body = models.TextField(verbose_name='正文')
    pub_date = models.DateTimeField(verbose_name='文章日期')

    created = models.DateTimeField(verbose_name='发表日志', auto_now_add=True)
    updated = models.DateTimeField(verbose_name='最后编辑', auto_now=True)

    class Meta:
        ordering = '-created',
        verbose_name = '文章'
        verbose_name_plural = verbose_name

    def __str__(self):
        return self.title


class Province(models.Model):
    name = models.CharField(verbose_name='省份名称', null=False, blank=False, max_length=64, unique=True)

    class Meta:
        ordering = 'name',
        verbose_name = '省份'
        verbose_name_plural = verbose_name

    def __str__(self):
        return self.name


class Area(models.Model):
    province = models.ForeignKey(
        to=Province, null=False, blank=False, on_delete=models.CASCADE,
        verbose_name='所属省份', related_name='area_ls')
    name = models.CharField(
        verbose_name='区域名称', max_length=64, null=False, blank=False, unique=False,
        )

    class Meta:
        ordering = 'province', 'name'
        verbose_name = '地区'
        verbose_name_plural = verbose_name

    def __str__(self):
        return f'{self.province.name},{self.name}'


HEADERS = [
    "ID",
    "省份",
    "地区",
    "年度",
    "规模以上工业企业R&D经费支出",
    "单位GDP能耗",
    "乡-城人均年收入比",
    "城镇居民人均可支配收入",
    "农村居民人均可支配收入",
    "生活垃圾无害化处理率",
    "平均万人拥有公共汽车",
    "城镇生活污水集中处理率",
    "平均受教育年限",
    "死亡率",
    "养老保险覆盖率",
    "医疗保险覆盖率",
    "失业保险覆盖率",
    "PM2.5年平均浓度",
    "二氧化硫排放量",
    "单位GDP二氧化碳排放量",
    "化学需氧量排放量",
    "氨氮排放量",
    "人均耗水量",
    "单位GDP用水量",
    "播种面积占比",
    "人均生态足迹"
           ]


def can_not_equal_zero(data):
    '''基础数据每一项都都必须是不能等于 0 的'''
    if data == 0:
        raise ValidationError('该项数据不能等于0')
    return True


def import_excel_validator(excel_file_obj, province_obj=None):
    try:
        wb = load_workbook(excel_file_obj, read_only=True)
    except:
        raise ValidationError('非法excel文件')
    ws = wb[wb.sheetnames[0]]
    if ws.max_column != len(HEADERS):
        raise ValidationError(f'数据应该包含: {"、".join(HEADERS)}，共计{len(HEADERS)}个字段.')
    provinces = [p.name for p in Province.objects.all()]
    r = 0
    province_name = ''
    year = ''
    cities = []
    the_cities = []
    for row in ws.iter_rows():  # 逐行验证数据
        r += 1
        if r == 1:
            c = 0
            _ = []
            for cell in row:
                c += 1
                if cell.value != HEADERS[c-1]:
                    _.append(f'第{c}列表头错误，应该是{HEADERS[c-1]}， 而实际是{cell.value}')
            if _:
                raise ValidationError(f'数据校验未通过：{";".join(_)}。')  # 验证数据列

        else:
            if r == 2:
                province_name = row[1].value  # 获取省份
                if province_name == "某某省":
                    return True
                if province_name not in provinces:  # 验证省份
                    raise ValidationError(f'第 {r} 行省份“{row[1].value}”在数据库中没有记录，请确保省份和地区的数据正确。')
                if isinstance(province_obj, Province):
                    if province_name != province_obj.name:
                        raise ValidationError(f'Excel中期望的是{province_obj.name}的数据，而实际是{province_name}的数据。')
                    else:
                        return True
                year = row[3].value
                if not isinstance(year, int) or len(str(year)) != 4:
                    raise ValidationError(f'第 {r} 行年份数据不正确，应为 4 位数字，比如 2017。')
                cities = [city.name for city in Area.objects.filter(province__name=province_name).all()]  # 获取所有省份下的地区
            _year = row[3].value  # 获取年份
            if _year != year:
                raise ValidationError(f'第 {r} 行年份数据不正确，应为 {year}，而实际上是 {_year}。')
            _city = row[2].value  # 获取地区
            if _city in the_cities:  # 验证不能重复
                raise ValidationError(f'第 {r} 行地区“{_city}”数据重复，省内每个地区的数据只能出现一次')
            else:
                the_cities.append(_city)
            if _city not in cities:  # 验证地区在合法范围内取值
                raise ValidationError(f'第 {r} 行地区数据不正确，应为 {"、".join(cities)}中的一个，而实际上是 “{_city}”。')
            _province_name = row[1].value
            if _province_name != province_name:  # 验证省份
                raise ValidationError(f'第 {r} 行地区数据不正确，应为 {province_name}中的一个，而实际上是 “{_province_name}”。')
            c = 0
            for cell in row[4:]:
                c += 1
                if not isinstance(cell.value, (float, int)):
                    raise ValidationError(f'第 {r} 行第{c}列({HEADERS[c-1]})数据格式不正确， 应为数值型。')
                if cell.value == 0:  # 确保每项数据都不是 0 fix 20200314
                    raise ValidationError(f'第 {r} 行第{c}列({HEADERS[c - 1]})数值为0，这个是不允许的。')
    missing_cities = []
    for city in cities:
        if city not in the_cities:
            missing_cities.append(city)
    if missing_cities:
        raise ValidationError(f'{"、".join(missing_cities)}地区的数据缺失')  # 验证缺失
    return True


class CityDataRecord(models.Model):

    # province_data = models.ForeignKey(to='ProvinceDataRecord', on_delete=models.CASCADE, related_name='cities_records', null=True, blank=True)

    area = models.ForeignKey(
        to=Area, null=False, blank=False, on_delete=models.CASCADE,
        verbose_name='地区', related_name='data_record_ls')

    year = models.IntegerField(verbose_name='年度', null=False, blank=False)

    r_d = models.FloatField(verbose_name='规模以上工业企业R&D经费支出(万元)', null=False, blank=False, validators=[can_not_equal_zero])
    per_unit_gdp = models.FloatField(verbose_name='单位GDP能耗(吨/万元)', null=False, blank=False, validators=[can_not_equal_zero])
    rural_urban = models.FloatField(verbose_name='乡-城人均年收入比(/)', null=False, blank=False, validators=[can_not_equal_zero])
    urban_per = models.FloatField(verbose_name='城镇居民人均可支配收入(元)', null=False, blank=False, validators=[can_not_equal_zero])
    rural_per = models.FloatField(verbose_name='农村居民人均可支配收入(元)', null=False, blank=False, validators=[can_not_equal_zero])
    garbage = models.FloatField(verbose_name='生活垃圾无害化处理率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    bus_per = models.FloatField(verbose_name='平均万人拥有公共汽车(/)', null=False, blank=False, validators=[can_not_equal_zero])
    urban_sewage = models.FloatField(verbose_name='城镇生活污水集中处理率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    edu_years = models.FloatField(verbose_name='平均受教育年限(年)', null=False, blank=False, validators=[can_not_equal_zero])

    mortality = models.FloatField(verbose_name='死亡率(%)', null=False, blank=False, validators=[can_not_equal_zero])

    pension_cov = models.FloatField(verbose_name='养老保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    medical_cov = models.FloatField(verbose_name='医疗保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    unemployment_cov = models.FloatField(verbose_name='失业保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])

    pm25 = models.FloatField(verbose_name='PM2.5年平均浓度(微克/立方米)', null=False, blank=False, validators=[can_not_equal_zero])
    so2_emissions = models.FloatField(verbose_name='二氧化硫排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    co2_per_gdp = models.FloatField(verbose_name='单位GDP二氧化碳排放量(千克/元)', null=False, blank=False, validators=[can_not_equal_zero])
    cod_emissions = models.FloatField(verbose_name='化学需氧量排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    nh_emissions = models.FloatField(verbose_name='氨氮排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    water_per = models.FloatField(verbose_name='人均耗水量(立方米)', null=False, blank=False, validators=[can_not_equal_zero])
    water_per_gdp = models.FloatField(verbose_name='单位GDP用水量(立方米/万元)', null=False, blank=False, validators=[can_not_equal_zero])

    planting_area = models.FloatField(verbose_name='播种面积占比(%)', null=False, blank=False, validators=[can_not_equal_zero])
    ef_per = models.FloatField(verbose_name='人均生态足迹(万吨碳/万人)', null=False, blank=False, validators=[can_not_equal_zero])

    # 后续计算
    green_innovation = models.FloatField(verbose_name='绿色创新', null=True, blank=True)
    energy_use = models.FloatField(verbose_name='能源利用', null=True, blank=True)
    parma_ratio = models.FloatField(verbose_name='帕尔玛比率', null=True, blank=True)
    income = models.FloatField(verbose_name='收入', null=True, blank=True)
    infrastructure = models.FloatField(verbose_name='基础设施建设', null=True, blank=True)

    education = models.FloatField(verbose_name='教育', null=True, blank=True)
    life_expectancy = models.FloatField(verbose_name='预期寿命', null=True, blank=True)
    social_security = models.FloatField(verbose_name='社会保障', null=True, blank=True)
    air_pollution = models.FloatField(verbose_name='大气污染', null=True, blank=True)
    greenhouse = models.FloatField(verbose_name='温室气体排放', null=True, blank=True)
    nitrogen = models.FloatField(verbose_name='氮排放', null=True, blank=True)
    water_withdrawal = models.FloatField(verbose_name='取水量', null=True, blank=True)
    land_use = models.FloatField(verbose_name='土地利用', null=True, blank=True)
    EF = models.FloatField(verbose_name='生态足迹', null=True, blank=True)

    city_green_economy = models.FloatField(verbose_name='绿色经济', null=True, blank=True)
    city_sustainable = models.FloatField(verbose_name='可持续发展', null=True, blank=True)
    city_gep_plus = models.FloatField(verbose_name='GEP+', null=True, blank=True)

    class Meta:
        ordering = '-year', 'area'
        verbose_name = '地区数据'
        verbose_name_plural = verbose_name
        unique_together = [('area', 'year')]

    def __str__(self):
        return f'{self.area}, {self.year}'

    def get_same_period_max_and_min(self, name):
        pass

    def get_last_year_value_and_step(self, name):
        pass

    def get_values(self, name):
        '''获取计算所需的值'''
        last_one = CityDataRecord.objects.filter(area=self.area, year=self.year - 1).first()
        # 前期值
        last_value = last_one.__dict__[name]
        # 当前值
        present_value = self.__dict__[name]

        # 同期最大

        # 同期最小

        # 变化值

    def get_last_year(self):
        return self.__class__.objects.filter(area=self.area, year=self.year-1).first()


class ProvinceDataRecord(models.Model):

    province = models.ForeignKey(
        to=Province, null=False, blank=False, on_delete=models.CASCADE,
        verbose_name='省份', related_name='data_record_ls')

    year = models.IntegerField(verbose_name='年度', null=False, blank=False)

    # 基础值
    r_d = models.FloatField(verbose_name='企业 R&D 内部经费支出(亿元)', null=False, blank=False, validators=[can_not_equal_zero])
    renewable_energy_per = models.FloatField(verbose_name='可再生能源供给占比(%)', null=False, blank=False, validators=[can_not_equal_zero])
    per_unit_gdp = models.FloatField(verbose_name='单位GDP能耗(吨/万元)', null=False, blank=False, validators=[can_not_equal_zero])
    rural_urban = models.FloatField(verbose_name='乡-城人均年收入比(/)', null=False, blank=False, validators=[can_not_equal_zero])
    urban_per = models.FloatField(verbose_name='城镇居民人均可支配收入(元)', null=False, blank=False, validators=[can_not_equal_zero])
    rural_per = models.FloatField(verbose_name='农村居民人均可支配收入(元)', null=False, blank=False, validators=[can_not_equal_zero])
    garbage = models.FloatField(verbose_name='生活垃圾无害化处理率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    bus_per = models.FloatField(verbose_name='平均万人拥有公共汽车(/)', null=False, blank=False, validators=[can_not_equal_zero])
    urban_sewage = models.FloatField(verbose_name='城镇生活污水集中处理率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    edu_years = models.FloatField(verbose_name='平均受教育年限(年)', null=False, blank=False, validators=[can_not_equal_zero])
    mortality = models.FloatField(verbose_name='死亡率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    pension_cov = models.FloatField(verbose_name='养老保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    medical_cov = models.FloatField(verbose_name='医疗保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    unemployment_cov = models.FloatField(verbose_name='失业保险覆盖率(%)', null=False, blank=False, validators=[can_not_equal_zero])
    pm25 = models.FloatField(verbose_name='PM2.5年平均浓度(微克/立方米)', null=False, blank=False, validators=[can_not_equal_zero])
    so2_emissions = models.FloatField(verbose_name='二氧化硫排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    co2_per_gdp = models.FloatField(verbose_name='单位GDP二氧化碳排放量(千克/元)', null=False, blank=False, validators=[can_not_equal_zero])
    cod_emissions = models.FloatField(verbose_name='化学需氧量排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    nh_emissions = models.FloatField(verbose_name='氨氮排放量(万吨)', null=False, blank=False, validators=[can_not_equal_zero])
    water_per = models.FloatField(verbose_name='人均耗水量(立方米)', null=False, blank=False, validators=[can_not_equal_zero])
    water_per_gdp = models.FloatField(verbose_name='单位GDP用水量(立方米/万元)', null=False, blank=False, validators=[can_not_equal_zero])
    planting_area = models.FloatField(verbose_name='播种面积占比(%)', null=False, blank=False, validators=[can_not_equal_zero])
    ef_per = models.FloatField(verbose_name='人均生态足迹(万吨碳/万人)', null=False, blank=False, validators=[can_not_equal_zero])

    # 目标值
    r_d_target = models.FloatField(verbose_name='企业 R&D 内部经费支出的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    renewable_energy_per_target = models.FloatField(verbose_name='可再生能源供给占比的目标值', null=False, blank=False,
                                             validators=[can_not_equal_zero])
    per_unit_gdp_target = models.FloatField(verbose_name='单位GDP能耗的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    rural_urban_target = models.FloatField(verbose_name='乡-城人均年收入比的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    urban_per_target = models.FloatField(verbose_name='城镇居民人均可支配收入的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    rural_per_target = models.FloatField(verbose_name='农村居民人均可支配收入的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    garbage_target = models.FloatField(verbose_name='生活垃圾无害化处理率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    bus_per_target = models.FloatField(verbose_name='平均万人拥有公共汽车的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    urban_sewage_target = models.FloatField(verbose_name='城镇生活污水集中处理率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    edu_years_target = models.FloatField(verbose_name='平均受教育年限的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    mortality_target = models.FloatField(verbose_name='死亡率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    pension_cov_target = models.FloatField(verbose_name='养老保险覆盖率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    medical_cov_target = models.FloatField(verbose_name='医疗保险覆盖率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    unemployment_cov_target = models.FloatField(verbose_name='失业保险覆盖率的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    pm25_target = models.FloatField(verbose_name='PM2.5年平均浓度的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    so2_emissions_target = models.FloatField(verbose_name='二氧化硫排放量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    co2_per_gdp_target = models.FloatField(verbose_name='单位GDP二氧化碳排放量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    cod_emissions_target = models.FloatField(verbose_name='化学需氧量排放量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    nh_emissions_target = models.FloatField(verbose_name='氨氮排放量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    water_per_target = models.FloatField(verbose_name='人均耗水量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    water_per_gdp_target = models.FloatField(verbose_name='单位GDP用水量的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    planting_area_target = models.FloatField(verbose_name='播种面积占比的目标值', null=False, blank=False, validators=[can_not_equal_zero])
    ef_per_target = models.FloatField(verbose_name='人均生态足迹的目标值', null=False, blank=False, validators=[can_not_equal_zero])

    # # attach
    #
    # attach = models.FileField(
    #     verbose_name='年度省内各地区数据', upload_to='data/%Y-%m-%d', default='data/数据模板.xlsx',
    #     validators=[
    #         FileExtensionValidator(allowed_extensions=('xlsx',), message='仅支持上传扩展名为 xlsx 的Excel文件'),
    #         import_excel_validator,
    #                 ],
    #                           )

    # 后续计算
    green_innovation = models.FloatField(verbose_name='绿色创新', null=True, blank=True)
    renewable_energy = models.FloatField(verbose_name='可再生能源供给', null=True, blank=True,
                                             validators=[can_not_equal_zero])
    energy_use = models.FloatField(verbose_name='能源利用', null=True, blank=True)
    parma_ratio = models.FloatField(verbose_name='帕尔玛比率', null=True, blank=True)
    income = models.FloatField(verbose_name='收入', null=True, blank=True)
    infrastructure = models.FloatField(verbose_name='基础设施建设', null=True, blank=True)

    education = models.FloatField(verbose_name='教育', null=True, blank=True)
    life_expectancy = models.FloatField(verbose_name='预期寿命', null=True, blank=True)
    social_security = models.FloatField(verbose_name='社会保障', null=True, blank=True)
    air_pollution = models.FloatField(verbose_name='大气污染', null=True, blank=True)
    greenhouse = models.FloatField(verbose_name='温室气体排放', null=True, blank=True)
    nitrogen = models.FloatField(verbose_name='氮排放', null=True, blank=True)
    water_withdrawal = models.FloatField(verbose_name='取水量', null=True, blank=True)
    land_use = models.FloatField(verbose_name='土地利用', null=True, blank=True)
    EF = models.FloatField(verbose_name='生态足迹', null=True, blank=True)

    city_green_economy = models.FloatField(verbose_name='绿色经济', null=True, blank=True)
    city_sustainable = models.FloatField(verbose_name='可持续发展', null=True, blank=True)
    city_gep_plus = models.FloatField(verbose_name='GEP+', null=True, blank=True)

    # 1 新上传 2 excel 地区数据解析入库成功 3 计算中 4 计算失败 5 计算成功 6 更新数据需要重新计算
    status = models.SmallIntegerField(verbose_name='状态', default=1, )

    class Meta:
        ordering = '-year', 'province'
        verbose_name = '省份数据'
        verbose_name_plural = verbose_name

    def __str__(self):
        return f'{self.province.name},{self.year}'




